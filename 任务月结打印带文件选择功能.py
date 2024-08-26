import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins
import tkinter as tk
from tkinter import filedialog
import threading

def process_excel(input_file, output_file):
    # Step 1: Read the data
    df = pd.read_csv(input_file)

    # Rename columns if necessary to match expected names
    df.rename(columns={'护理对象姓名': '护理对象', '长护服务项目': '服务项目', '护理日期': '护理日期',
                       '护理人员': '护理人员'}, inplace=True)

    # Step 2: Convert the relevant columns to datetime
    df['护理日期'] = pd.to_datetime(df['护理日期'])

    # Step 3: Group by '护理对象', '护理日期', and '服务项目', and count the number of occurrences of each service project
    detailed_grouped_df = df.groupby(['护理对象', df['护理日期'].dt.date, '服务项目']).size().reset_index(
        name='项目次数')

    # Step 4: Create a pivot table
    pivot_table = pd.pivot_table(
        detailed_grouped_df,
        values='项目次数',
        index=['护理对象', '服务项目'],
        columns='护理日期',
        aggfunc='sum',
        fill_value=0
    )

    # Step 5: Update the column names to only include the day part of the date
    pivot_table.columns = [col.day for col in pivot_table.columns]

    # Step 6: Replace 0 with empty strings
    pivot_table = pivot_table.applymap(lambda x: '' if x == 0 else x)

    # Step 7: Add a new column for "项目小计" which sums each row
    pivot_table['项目小计'] = pivot_table.apply(lambda row: pd.to_numeric(row, errors='coerce').sum(), axis=1)

    # Step 8: Export the updated pivot table to an Excel file
    pivot_table.reset_index(inplace=True)  # Reset index to ensure '护理对象' and '服务项目' are columns
    pivot_table.to_excel(output_file, sheet_name='数据透视表', index=False)

    # Step 9: Load the workbook and select the active worksheet to apply borders and add title row
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # ... (您的原有代码)
    # Add the new title row
    title = f"{df['护理人员'].iloc[0]} {df['护理日期'].dt.strftime('%Y-%m').iloc[0]} 护理任务确认表"
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
    ws.cell(row=1, column=1).value = title
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # Define the border style
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    # Apply the border to all cells in the worksheet starting from the second row
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border

    # Merge cells for the same "护理对象" and center align
    current_person = None
    start_row = 2  # Start from the second row (first row is the title)
    for row in range(2, ws.max_row + 1):
        person = ws.cell(row=row, column=1).value
        if person != current_person:
            if current_person is not None:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                ws.cell(start_row, 1).alignment = Alignment(horizontal='center', vertical='center')
            current_person = person
            start_row = row
        elif row == ws.max_row:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=row, end_column=1)
            ws.cell(start_row, 1).alignment = Alignment(horizontal='center', vertical='center')

    # Set page setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # Set header and footer
    ws.oddHeader.center.text = "青田唯康养老服务有限公司长护险居家上门护理任务确认表"
    ws.oddFooter.center.text = "页码 &P / &N          护理员签名："

    # Set page margins
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=1, bottom=1, header=0.5, footer=0.5)
    # Save the workbook with the updated borders and new column
    wb.save(output_file)
    import time
    time.sleep(5)  # 模拟耗时操作
    print(f"Processed {input_file} and saved to {output_file}")

def select_files():
    input_file = filedialog.askopenfilename(title="选择输入文件")
    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", title="选择输出文件")
    if input_file and output_file:
        # 创建并启动一个新线程来处理文件
        threading.Thread(target=process_excel, args=(input_file, output_file)).start()

# 创建主窗口
root = tk.Tk()
root.title("任务统计")

# 创建选择文件按钮
button = tk.Button(root, text="选择文件", command=select_files)
button.pack()

# 启动GUI程序
root.mainloop()
